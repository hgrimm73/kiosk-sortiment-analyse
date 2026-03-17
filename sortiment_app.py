import os
import streamlit as st
import pandas as pd
import re
import io
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

st.set_page_config(page_title="Kiosk Sortiment Analyse", page_icon="🏟️", layout="wide")


# ─────────────────────────────────────────────
# 1. PASSWORT-SCHUTZ
# ─────────────────────────────────────────────
def check_password():
    if "password_correct" not in st.session_state:
        st.markdown(
            "<h1 style='text-align:center;'>🔐 Interner Login</h1>",
            unsafe_allow_html=True,
        )
        pwd = st.text_input("Passwort eingeben:", type="password", key="init_pw")
        if pwd == "makeitso!":
            st.session_state["password_correct"] = True
            st.rerun()
        elif pwd != "":
            st.error("Passwort falsch.")
        return False
    return True


# ─────────────────────────────────────────────
# 2. HILFSFUNKTIONEN
# ─────────────────────────────────────────────
def normalize(s):
    if not s or pd.isna(s):
        return ""
    return " ".join(str(s).replace("\n", " ").split()).strip()


def format_k_list(ks):
    if not ks:
        return "-"
    nums = sorted(list(set([
        int(re.search(r"\d+", str(k)).group())
        for k in ks if re.search(r"\d+", str(k))
    ])))
    return "K" + "-".join([str(n).zfill(2) for n in nums])


def checked_filename(original_name):
    """
    'foo.xlsx' -> 'foo_checked.xlsx'
    'foo.pdf'  -> 'foo_checked.xlsx'  (PDF ist nicht editierbar; Korrekturen
                                        werden als Excel-Datei gespeichert)
    """
    base, ext = os.path.splitext(original_name)
    out_ext = ".xlsx" if ext.lower() == ".pdf" else ext
    return base + "_checked" + out_ext


def file_type(name):
    """Returns 'pdf', 'xlsx', or 'unknown'."""
    n = name.lower()
    if n.endswith(".pdf"):
        return "pdf"
    if n.endswith(".xlsx") or n.endswith(".xls"):
        return "xlsx"
    return "unknown"


# ─────────────────────────────────────────────
# 3. DATEI EINLESEN (PDF + Excel unified)
# ─────────────────────────────────────────────
def extract_tables_from_pdf(file_bytes):
    """
    Liest alle Tabellen aus PDF-Bytes.
    Gleich breite Sub-Tabellen einer Seite werden horizontal zusammengefuehrt.
    """
    try:
        all_page_rows = None
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    continue
                row_counts = [len(t) for t in tables]
                if len(tables) > 1 and len(set(row_counts)) == 1:
                    n_rows = row_counts[0]
                    page_rows = []
                    for r in range(n_rows):
                        row = []
                        for t in tables:
                            row.extend([
                                str(c).strip().replace("\n", " ") if c is not None else ""
                                for c in t[r]
                            ])
                        page_rows.append(row)
                else:
                    page_rows = []
                    for t in tables:
                        for row in t:
                            page_rows.append([
                                str(c).strip().replace("\n", " ") if c is not None else ""
                                for c in row
                            ])
                all_page_rows = (all_page_rows or []) + page_rows

        if not all_page_rows:
            return None, "Keine Tabellen im PDF gefunden."

        max_cols = max(len(r) for r in all_page_rows)
        df = pd.DataFrame([r + [""] * (max_cols - len(r)) for r in all_page_rows])
        return df, None
    except Exception as e:
        return None, "Fehler beim Lesen der PDF: " + str(e)


def load_file_df(uploaded_file):
    """
    Laedt eine Datei (PDF oder Excel) und gibt (DataFrame, Bytes, Fehler) zurueck.
    Bytes werden fuer den _checked-Download gespeichert.
    """
    try:
        uploaded_file.seek(0)
        raw_bytes = uploaded_file.read()
        ftype = file_type(uploaded_file.name)

        if ftype == "pdf":
            df, err = extract_tables_from_pdf(raw_bytes)
            return df, raw_bytes, err

        elif ftype == "xlsx":
            df = pd.read_excel(io.BytesIO(raw_bytes), header=None).fillna("").astype(str)
            return df, raw_bytes, None

        else:
            return None, raw_bytes, "Nicht unterstuetzter Datei-Typ (nur .pdf und .xlsx)."

    except Exception as e:
        return None, b"", str(e)


def make_checked_bytes(edited_df, original_bytes, ftype):
    """
    Erstellt die _checked-Datei immer als Excel.
    PDFs sind nicht rueckschreibbar; die Korrekturen aus dem Editor
    existieren nur im DataFrame. Daher wird in beiden Faellen
    das editierte DataFrame als .xlsx exportiert.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        edited_df.to_excel(writer, index=False, header=False)
    return output.getvalue()


# ─────────────────────────────────────────────
# 4. ANOMALIE-ERKENNUNG
# ─────────────────────────────────────────────
def _get_header_info(df):
    """Gibt (h_row, k_cols, n_col, p_col) zurueck."""
    h_row, k_cols, n_col, p_col = -1, [], 1, 5
    for i, row in df.iterrows():
        row_vals = [normalize(str(x)).upper() for x in row]
        kiosk_hits = [
            c for c, v in enumerate(row_vals)
            if re.search(r"KIOSK.*\d", v.replace(" ", ""))
        ]
        if len(kiosk_hits) >= 2:
            h_row = i
            k_cols = kiosk_hits
            for c, v in enumerate(row_vals):
                if any(x in v for x in ["PRODUKT", "ARTIKEL", "BEZEICHNUNG"]):
                    n_col = c
                if any(x in v for x in ["PREIS", "VK PREIS", "BRUTTO"]):
                    p_col = c
            break
    return h_row, k_cols, n_col, p_col


def detect_row_anomalies(df):
    """
    Erkennt:
      'category_with_x'   – Kategorie-Kopfzeile (kein Preis) hat ein X in Kiosk-Spalte
      'product_empty_kiosk' – Produkt (hat Preis) hat leere Kiosk-Felder (weder X noch -)
    """
    h_row, k_cols, n_col, p_col = _get_header_info(df)
    if h_row == -1:
        return {}

    header = df.iloc[h_row]
    k_label = {c: normalize(str(header.iloc[c])) for c in k_cols}
    anomalies = {}

    for i, row in df.iloc[h_row + 1:].iterrows():
        row_vals_up = [normalize(str(x)).upper() for x in row]
        if sum(1 for v in row_vals_up
               if re.search(r"KIOSK.*\d", v.replace(" ", ""))) >= 2:
            continue

        name  = normalize(str(row.iloc[n_col]) if n_col < len(row) else "")
        price = normalize(str(row.iloc[p_col]) if p_col < len(row) else "")
        if not name:
            continue

        kiosk_vals = {c: str(row.iloc[c]).strip() if c < len(row) else "" for c in k_cols}

        if not price:
            x_cols = [c for c, v in kiosk_vals.items() if v.upper() == "X"]
            if x_cols:
                anomalies[i] = {
                    "type": "category_with_x",
                    "name": name,
                    "x_cols": x_cols,
                    "empty_cols": [],
                    "kiosk_labels": [k_label.get(c, str(c)) for c in x_cols],
                }

        if price:
            empty_cols = [c for c, v in kiosk_vals.items() if v not in ("X", "x", "-")]
            if empty_cols:
                anomalies[i] = {
                    "type": "product_empty_kiosk",
                    "name": name,
                    "x_cols": [],
                    "empty_cols": empty_cols,
                    "kiosk_labels": [k_label.get(c, str(c)) for c in empty_cols],
                }

    return anomalies


def detect_issues(df):
    """Gibt Liste von (typ, nachricht)-Tupeln zurueck fuer den Issue-Expander."""
    issues = []
    h_row, k_cols, n_col, _ = _get_header_info(df)

    if h_row == -1:
        issues.append(("error", "Keine Kopfzeile mit Kiosk-Spalten gefunden."))
        return issues

    repeated = [
        i + 1 for i, row in df.iloc[h_row + 1:].iterrows()
        if sum(1 for v in [normalize(str(x)).upper() for x in row]
               if re.search(r"KIOSK.*\d", v.replace(" ", ""))) >= 2
    ]
    if repeated:
        issues.append((
            "warning",
            str(len(repeated)) + " wiederholte Kopfzeile(n) erkannt "
            "(Seitenumbrueche) – werden automatisch ignoriert.",
        ))

    anomalies = detect_row_anomalies(df)
    cat_x  = [a for a in anomalies.values() if a["type"] == "category_with_x"]
    prod_e = [a for a in anomalies.values() if a["type"] == "product_empty_kiosk"]

    if cat_x:
        det = "; ".join(
            '"' + a["name"] + '" (bei ' + ", ".join(a["kiosk_labels"]) + ")"
            for a in cat_x[:3]
        ) + (" ..." if len(cat_x) > 3 else "")
        issues.append(("warning",
            str(len(cat_x)) + " Kategorie-Zeile(n) mit irrtaeglichem X: " + det))

    if prod_e:
        det = "; ".join(
            '"' + a["name"] + '" (' + str(len(a["empty_cols"])) + " leere Felder)"
            for a in prod_e[:3]
        ) + (" ..." if len(prod_e) > 3 else "")
        issues.append(("info",
            str(len(prod_e)) + " Produkt(e) mit leeren Kiosk-Feldern: " + det))

    if not issues:
        issues.append(("success", "Keine Auffaelligkeiten – Daten sehen gut aus."))

    return issues


def style_raw_df(df, anomalies):
    """Faerbt Anomalie-Zellen ein (gelb/orange fuer Kategorien, rot fuer leere Felder)."""
    col_names = list(df.columns)
    style_map = pd.DataFrame("", index=df.index, columns=df.columns)

    for row_idx, info in anomalies.items():
        if row_idx not in df.index:
            continue
        if info["type"] == "category_with_x":
            style_map.loc[row_idx, :] = "background-color: #fff3cd;"
            for c in info["x_cols"]:
                if c < len(col_names):
                    style_map.loc[row_idx, col_names[c]] = (
                        "background-color: #fd7e14; color: white; font-weight: bold;"
                    )
        elif info["type"] == "product_empty_kiosk":
            for c in info["empty_cols"]:
                if c < len(col_names):
                    style_map.loc[row_idx, col_names[c]] = (
                        "background-color: #f8d7da; color: #721c24;"
                    )

    return df.style.apply(lambda _: style_map, axis=None)


# ─────────────────────────────────────────────
# 5. PARSING
# ─────────────────────────────────────────────
def parse_df_to_result(df, filename):
    try:
        df = pd.DataFrame(df).fillna("").astype(str)
        # Zeile-Spalte entfernen falls vom Review-Editor hinzugefuegt
        if "Zeile" in df.columns:
            df = df.drop(columns=["Zeile"])

        h_row, k_map = -1, {}
        n_col = p_col = w_col = None
        is_unit_col = False

        for i, row in df.iterrows():
            row_vals = [normalize(x).upper() for x in row]
            kiosk_hits = [
                c for c, v in enumerate(row_vals)
                if re.search(r"KIOSK.*\d", v.replace(" ", ""))
            ]
            if len(kiosk_hits) >= 2:
                h_row = i
                for c, v in enumerate(row_vals):
                    if re.search(r"KIOSK.*\d", v.replace(" ", "")):
                        num = re.search(r"\d+", v)
                        k_map[c] = "K" + (num.group().zfill(2) if num else str(c))
                    if any(x in v for x in ["PRODUKT", "ARTIKEL", "BEZEICHNUNG"]):
                        n_col = c
                    if any(x in v for x in ["PREIS", "VK PREIS", "BRUTTO"]):
                        p_col = c
                    if any(x in v for x in ["WARENGRUPPE", "GRUPPE"]):
                        w_col = c; is_unit_col = False
                    if "EINHEIT" in v:
                        w_col = c; is_unit_col = True
                break

        if h_row == -1:
            return None

        kiosk_cols = set(k_map.keys())

        if n_col is None or n_col in kiosk_cols:
            cands = [c for c in range(df.shape[1]) if c not in kiosk_cols]
            if cands:
                text_cnt = {
                    c: sum(1 for v in df.iloc[h_row + 1:][c]
                           if normalize(str(v)) not in ("X", "x", "", "0", "-"))
                    for c in cands
                }
                n_col = max(cands, key=lambda c: text_cnt.get(c, 0))

        if p_col is None or p_col in kiosk_cols:
            cands = [c for c in range(df.shape[1])
                     if c not in kiosk_cols and c != n_col]
            price_re = re.compile(r"\d+[.,]\d{2}")
            best_p, best_s = (cands[0] if cands else (n_col or 0) + 1), 0
            for c in cands:
                s = sum(1 for v in df.iloc[h_row + 1:][c] if price_re.search(str(v)))
                if s > best_s:
                    best_s, best_p = s, c
            p_col = best_p

        if w_col is None or w_col in kiosk_cols:
            fb = [c for c in range(df.shape[1])
                  if c not in kiosk_cols and c != n_col and c != p_col]
            w_col = fb[0] if fb else n_col

        food, drinks = [], []
        sec, current_cat = "FOOD", "ALLGEMEIN"

        for i, row in df.iloc[h_row + 1:].iterrows():
            name_val  = normalize(row.iloc[n_col] if n_col < len(row) else "")
            price_val = normalize(row.iloc[p_col] if p_col < len(row) else "")
            cat_val   = normalize(row.iloc[w_col] if w_col < len(row) else "")
            clean_name = name_val.upper()

            if clean_name in ["GETRAENKE", "GETRANKE", "GETRÄNKE", "DRINKS"]:
                sec = "DRINKS"; current_cat = "GETRÄNKE"; continue
            if clean_name in ["FOOD", "SPEISEN"]:
                sec = "FOOD"; current_cat = "FOOD"; continue
            if sum(1 for v in [normalize(str(x)).upper() for x in row]
                   if re.search(r"KIOSK.*\d", v.replace(" ", ""))) >= 2:
                continue

            marked_ks = [
                k_map[c] for c in k_map
                if c < len(row) and str(row.iloc[c]).strip().upper() == "X"
            ]
            is_product = price_val != "" or len(marked_ks) > 0

            if not is_product and name_val:
                current_cat = name_val; continue
            if is_product and not is_unit_col and cat_val:
                current_cat = cat_val
            if is_product and name_val:
                item = {"cat": current_cat, "name": name_val,
                        "price": price_val, "ks": marked_ks}
                (food if sec == "FOOD" else drinks).append(item)

        return {
            "food": food, "drinks": drinks,
            "ks": sorted(list(k_map.values())),
            "name": filename,
            "_cols": {"n": n_col, "p": p_col, "w": w_col, "kiosk": sorted(kiosk_cols)},
        }
    except Exception:
        return None


# ─────────────────────────────────────────────
# 6. EXPORTS
# ─────────────────────────────────────────────
def create_excel_export(data):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sortiment"
    ws.append(["Bereich", "Warengruppe", "Produkt", "Preis"])
    row_i = 2
    for label, items in [("FOOD", data["food"]), ("GETRÄNKE", data["drinks"])]:
        if not items:
            continue
        s_row = row_i
        groups = {}
        for k in data["ks"]:
            assort = tuple([(i["cat"], i["name"], i["price"])
                            for i in items if k in i["ks"]])
            if assort not in groups:
                groups[assort] = []
            groups[assort].append(k)
        for assort, ks in sorted(groups.items(), key=lambda x: x[1][0]):
            ws.merge_cells("B" + str(row_i) + ":D" + str(row_i))
            ws.cell(row=row_i, column=2,
                    value="Kioske: " + format_k_list(ks)).font = Font(bold=True)
            row_i += 1
            for cat, n, p in assort:
                ws.cell(row=row_i, column=2, value=cat)
                ws.cell(row=row_i, column=3, value=n)
                ws.cell(row=row_i, column=4, value=p)
                row_i += 1
            row_i += 1
        ws.merge_cells("A" + str(s_row) + ":A" + str(row_i - 2))
        ws.cell(row=s_row, column=1, value=label).font = Font(bold=True)
        row_i += 1
    wb.save(output)
    return output.getvalue()


def create_kiosk_diff_report(old, new):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Grafiker-Anweisung"
    ws.append(["Kiosk(e)", "Bereich", "Zugehoerigkeit Alt", "Zugehoerigkeit Neu",
               "Status", "Details der Aenderungen"])

    def get_kiosk_map(data):
        m = {}
        for skey in ["food", "drinks"]:
            assorts = {}
            for k in data["ks"]:
                asort = tuple(sorted([(i["name"], i["price"])
                                       for i in data[skey] if k in i["ks"]]))
                if asort not in assorts:
                    assorts[asort] = []
                assorts[asort].append(k)
            for asort, ks in assorts.items():
                name = format_k_list(ks)
                for k in ks:
                    m[(k, skey)] = (asort, name)
        return m

    old_m = get_kiosk_map(old)
    new_m = get_kiosk_map(new)
    report_groups = {}

    for skey, label in [("food", "FOOD"), ("drinks", "GETRÄNKE")]:
        old_groups = sorted(list(set([
            old_m.get((k, skey), (None, "Nicht vorhanden"))[1]
            for k in old["ks"]
        ])))
        for og_name in old_groups:
            if og_name == "Nicht vorhanden":
                continue
            ks_in_og = [k for k in old["ks"] if old_m.get((k, skey))[1] == og_name]
            sub_splits = {}
            for k in ks_in_og:
                n_asort, ng_name = new_m.get((k, skey), (tuple(), "Nicht vorhanden"))
                o_asort = old_m.get((k, skey))[0]
                if n_asort != o_asort:
                    if (n_asort, ng_name) not in sub_splits:
                        sub_splits[(n_asort, ng_name)] = []
                    sub_splits[(n_asort, ng_name)].append(k)
            for (n_asort, ng_name), ks_list in sub_splits.items():
                o_asort = old_m.get((ks_list[0], skey))[0]
                o_d, n_d = dict(o_asort), dict(n_asort)
                changes, status = [], ["Inhalt geaendert"]
                if og_name != ng_name:
                    status.append("Gruppe gewechselt / Split")
                for a in sorted(set(n_d) - set(o_d)):
                    changes.append("[+] Neu: " + a + " (" + n_d[a] + ")")
                for r in sorted(set(o_d) - set(n_d)):
                    changes.append("[-] Weg: " + r)
                for p in sorted(n for n in set(o_d) & set(n_d) if o_d[n] != n_d[n]):
                    changes.append("[!] Preis: " + p + " (" + o_d[p] + " -> " + n_d[p] + ")")
                group_key = (
                    format_k_list(ks_list), label, og_name, ng_name,
                    ", ".join(status), "\n".join(changes),
                )
                report_groups[group_key] = True

    row_i = 2
    for (k_ids, area, o_grp, n_grp, stat, det) in report_groups.keys():
        ws.cell(row=row_i, column=1, value=k_ids)
        ws.cell(row=row_i, column=2, value=area)
        ws.cell(row=row_i, column=3, value=o_grp)
        ws.cell(row=row_i, column=4, value=n_grp)
        ws.cell(row=row_i, column=5, value=stat)
        ws.cell(row=row_i, column=6, value=det).alignment = Alignment(wrap_text=True)
        row_i += 1
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["F"].width = 70
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# 7. ANALYSE-ANZEIGE
# ─────────────────────────────────────────────
def show_analysis_ui(res, source_filename, key_prefix=""):
    if "_cols" in res:
        c = res["_cols"]
        st.caption(
            "Erkannte Spalten — Produktname: " + str(c["n"]) +
            ", Preis: " + str(c["p"]) +
            ", Warengruppe: " + str(c["w"]) +
            ", Kiosk-Spalten: " + str(c["kiosk"])
        )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Wie viele Dateien benoetigt?", type="primary",
                     key=key_prefix + "_files_btn"):
            f_t, d_t = 0, 0
            for s, l in [("food", "**🍔 FOOD**"), ("drinks", "**🥤 GETRÄNKE**")]:
                st.markdown(l)
                grps = {}
                for k in res["ks"]:
                    asort = tuple(sorted([(i["name"], i["price"])
                                          for i in res[s] if k in i["ks"]]))
                    if asort:
                        if asort not in grps:
                            grps[asort] = []
                        grps[asort].append(k)
                for ks in grps.values():
                    st.code(format_k_list(ks), language=None)
                    if s == "food": f_t += 1
                    else:           d_t += 1
            st.divider()
            st.write("FOOD: " + str(f_t) + " | GETRÄNKE: " + str(d_t) +
                     " | **GESAMT: " + str(f_t + d_t) + "**")
    with col2:
        st.download_button(
            "📥 Excel Export",
            data=create_excel_export(res),
            file_name="Analyse_" + source_filename + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key_prefix + "_dl_btn",
        )

    for label, items in [("FOOD", res["food"]), ("GETRÄNKE", res["drinks"])]:
        st.subheader(label)
        grps = {}
        for k in res["ks"]:
            asort = tuple([(i["cat"], i["name"], i["price"])
                           for i in items if k in i["ks"]])
            if asort not in grps:
                grps[asort] = []
            grps[asort].append(k)
        for asort, ks in sorted(grps.items(), key=lambda x: x[1][0]):
            with st.expander("Kioske: " + format_k_list(ks)):
                curr = ""
                for cat, n, p in asort:
                    if cat != curr:
                        st.markdown("**" + cat + "**")
                        curr = cat
                    st.write("- " + n + ": " + p)


# ─────────────────────────────────────────────
# 8. REVIEW-SCHRITT (wiederverwendbar fuer Tab 1)
# ─────────────────────────────────────────────
def show_review_step(prefix, up_file):
    """
    Steuert den vollstaendigen Review-Workflow fuer eine Datei in Tab 1.
    Session-State-Keys werden mit 'prefix' getrennt, damit mehrere
    Instanzen nicht kollidieren.

    Gibt (result | None, confirmed: bool) zurueck.
    """
    fname    = up_file.name
    ftype_v  = file_type(fname)
    ss       = st.session_state

    # Neues File: State zuruecksetzen
    if ss.get(prefix + "_fname") != fname:
        for k in ("_fname", "_raw_df", "_raw_bytes", "_ftype",
                  "_confirmed", "_result", "_checked_bytes"):
            ss[prefix + k] = None
        ss[prefix + "_fname"] = fname

    # Einmalig laden
    if ss.get(prefix + "_raw_df") is None:
        with st.spinner("Datei wird eingelesen ..."):
            df, raw_bytes, err = load_file_df(up_file)
        if err:
            st.error(err)
            return None, False
        ss[prefix + "_raw_df"]   = df
        ss[prefix + "_raw_bytes"] = raw_bytes
        ss[prefix + "_ftype"]    = ftype_v

    raw_df    = ss[prefix + "_raw_df"]
    raw_bytes = ss[prefix + "_raw_bytes"]
    ftype_v   = ss[prefix + "_ftype"]

    # ── Phase A: Review ──────────────────────────────────────────
    if not ss.get(prefix + "_confirmed"):

        issues    = detect_issues(raw_df)
        anomalies = detect_row_anomalies(raw_df)
        has_warn  = any(t in ("error", "warning") for t, _ in issues)

        with st.expander(
            "🔍 Auffaelligkeiten" +
            (" – Bitte pruefen!" if has_warn else " – Alles ok"),
            expanded=has_warn,
        ):
            for itype, msg in issues:
                if itype == "error":    st.error(msg)
                elif itype == "warning": st.warning(msg)
                elif itype == "info":    st.info(msg)
                else:                   st.success(msg)

        # Farbige Vorschau (nur wenn Anomalien vorhanden)
        if anomalies:
            st.markdown(
                "**Farbmarkierungen:**  "
                "🟡 Gelb = Kategorie-Zeile mit irrtaeglichem X &nbsp;|&nbsp; "
                "🟠 Orange = fehlerhafte Zelle &nbsp;|&nbsp; "
                "🔴 Rot = leere Kiosk-Zelle (weder X noch -)",
                unsafe_allow_html=True,
            )
            styled = style_raw_df(raw_df, anomalies)
            st.dataframe(styled, use_container_width=True, height=350)

        # Bearbeitbarer Editor  ← PUNKT 1: explizite Zeile-Spalte
        st.markdown(
            "**Rohdaten korrigieren** – direkt bearbeiten, dann bestaetigen."
            "  Die Spalte **Zeile** entspricht der Zeilennummer in der Vorschau oben."
        )

        # Zeile-Spalte vorne einfuegen (read-only durch disabled=[])
        editor_df = raw_df.copy()
        editor_df.insert(0, "Zeile", list(raw_df.index))

        edited_with_zeile = st.data_editor(
            editor_df,
            use_container_width=True,
            num_rows="dynamic",
            disabled=["Zeile"],
            key=prefix + "_editor",
        )

        col_btn, col_hint = st.columns([1, 3])
        with col_btn:
            confirm = st.button(
                "Bestaetigen & Analysieren",
                type="primary",
                key=prefix + "_confirm_btn",
            )
        with col_hint:
            st.caption(
                "Kiosk-Zuordnungen = 'X' | Bereichs-Trenner: Zeile mit nur 'FOOD' / 'GETRÄNKE'"
            )

        if confirm:
            # Zeile-Spalte vor dem Parsen entfernen
            edited_df = edited_with_zeile.drop(columns=["Zeile"], errors="ignore")
            with st.spinner("Analyse laeuft ..."):
                parsed = parse_df_to_result(edited_df, fname)
            if parsed:
                checked = make_checked_bytes(edited_df, raw_bytes, ftype_v)
                ss[prefix + "_result"]        = parsed
                ss[prefix + "_confirmed"]     = True
                ss[prefix + "_checked_bytes"] = checked
                st.rerun()
            else:
                st.error("Analyse fehlgeschlagen – bitte Kiosk-Kopfzeile pruefen.")

        return None, False

    # ── Phase B: Analyse-Anzeige ─────────────────────────────────
    res    = ss[prefix + "_result"]
    checked_bytes = ss.get(prefix + "_checked_bytes", b"")

    st.success(
        "Datei importiert – " +
        str(len(res["food"])) + " Food- und " +
        str(len(res["drinks"])) + " Getränke-Produkte erkannt."
    )

    # PUNKT 2: _checked Download
    col_back, col_dl = st.columns([1, 2])
    with col_back:
        if st.button("Zurueck zur Datenpruefung", key=prefix + "_back_btn"):
            ss[prefix + "_confirmed"] = False
            st.rerun()
    with col_dl:
        c_fname = checked_filename(fname)
        c_mime  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        label   = ("💾 Als Excel speichern (" + c_fname + ")"
                   if ftype_v == "pdf"
                   else "💾 Geprueft speichern (" + c_fname + ")")
        st.download_button(
            label,
            data=checked_bytes,
            file_name=c_fname,
            mime=c_mime,
            key=prefix + "_checked_dl",
        )

    st.divider()
    show_analysis_ui(res, fname, key_prefix=prefix + "_ana")
    return res, True


# ─────────────────────────────────────────────
# 9. SCHNELL-PARSE fuer Tab 2 (kein Review)
# ─────────────────────────────────────────────
def quick_parse_file(up_file):
    """Laedt und parst eine Datei ohne Review-Schritt. Gibt (result, anomaly_count) zurueck."""
    df, _, err = load_file_df(up_file)
    if err or df is None:
        return None, 0, err or "Unbekannter Fehler"
    result = parse_df_to_result(df, up_file.name)
    if result is None:
        return None, 0, "Datei konnte nicht geparst werden (Kiosk-Kopfzeile fehlt?)"
    n_anom = len(detect_row_anomalies(df))
    return result, n_anom, None


# ─────────────────────────────────────────────
# 10. VERGLEICHS-ANZEIGE
# ─────────────────────────────────────────────
def show_diff_ui(old_res, new_res):
    b1, b2 = st.columns(2)
    with b1:
        do_anal = st.button("Unterschieds-Analyse starten", type="primary",
                            key="t2_analyse_btn")
    with b2:
        report_data = create_kiosk_diff_report(old_res, new_res)
        st.download_button(
            "📄 Unterschiede als Excel",
            data=report_data,
            file_name="Zusammenfassung_Grafiker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="t2_diff_dl",
        )

    if not do_anal:
        return

    for skey, title in [("food", "FOOD"), ("drinks", "GETRÄNKE")]:
        st.markdown("## " + title)
        o_grps = {}
        for k in old_res["ks"]:
            asort = tuple(sorted([
                (i["name"], i["price"]) for i in old_res[skey] if k in i["ks"]
            ]))
            if asort not in o_grps:
                o_grps[asort] = []
            o_grps[asort].append(k)

        for o_asort, o_ks in sorted(o_grps.items(), key=lambda x: x[1][0]):
            new_variants = {}
            for k in o_ks:
                n_asort = tuple(sorted([
                    (i["name"], i["price"]) for i in new_res[skey] if k in i["ks"]
                ]))
                if n_asort not in new_variants:
                    new_variants[n_asort] = []
                new_variants[n_asort].append(k)

            st.subheader("Ehemalige Gruppe: " + format_k_list(o_ks))
            if len(new_variants) == 1:
                n_asort = list(new_variants.keys())[0]
                lbl = "STABIL" if n_asort == o_asort else "GEAENDERT"
                st.markdown(
                    '<p class="status-stable">Status: ' + lbl + '</p>',
                    unsafe_allow_html=True,
                )
                o_d, n_d = dict(o_asort), dict(n_asort)
                for name in sorted(set(o_d) | set(n_d)):
                    if name not in o_d:      st.success("[+] " + name + ": " + n_d[name])
                    elif name not in n_d:    st.error("[-] " + name)
                    elif o_d[name] != n_d[name]:
                        st.warning("[!] " + name + ": " + o_d[name] + " -> " + n_d[name])
            else:
                st.markdown(
                    '<p class="status-split">Status: STRUKTURBRUCH / SPLIT</p>',
                    unsafe_allow_html=True,
                )
                for idx, (n_asort, sub_ks) in enumerate(new_variants.items()):
                    with st.expander(
                        "Untergruppe " + str(idx + 1) + ": " + format_k_list(sub_ks)
                    ):
                        o_d, n_d = dict(o_asort), dict(n_asort)
                        for name in sorted(set(o_d) | set(n_d)):
                            if name not in o_d:      st.success("[+] " + name + ": " + n_d[name])
                            elif name not in n_d:    st.error("[-] " + name)
                            elif o_d[name] != n_d[name]:
                                st.warning("[!] " + name + ": " + o_d[name] + " -> " + n_d[name])
            st.divider()


# ─────────────────────────────────────────────
# 11. HAUPT-UI
# ─────────────────────────────────────────────
if check_password():
    st.markdown("""
        <style>
        .main-title  { font-size: 2.2rem; font-weight: 700; margin-bottom: 1rem; }
        .status-stable { color: #0984e3; font-weight: bold;
                         border-left: 5px solid #0984e3; padding-left: 10px; }
        .status-split  { color: #d63031; font-weight: bold;
                         border-left: 5px solid #d63031; padding-left: 10px; }
        </style>
        <div class="main-title">🏟️ Analyse Verkaufssortimente – V 2.0</div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["1. Einzel-Analyse", "Verkaufssortimente-Vergleich"])

    # ════════════════════════════════════════
    # TAB 1 – EINZEL-ANALYSE
    # ════════════════════════════════════════
    with tab1:
        # PUNKT 3: kein Format-Radio mehr, auto-detect
        up = st.file_uploader(
            "Datei hochladen (PDF oder Excel)",
            type=["pdf", "xlsx"],
            key="t1_up",
        )
        if up:
            ftype_v = file_type(up.name)
            if ftype_v == "unknown":
                st.error("Nicht unterstuetzter Datei-Typ.")
            else:
                show_review_step("t1", up)

    # ════════════════════════════════════════
    # TAB 2 – VERGLEICH (PUNKT 4: alle Kombinationen)
    # ════════════════════════════════════════
    with tab2:
        st.header("Vergleich zwischen zwei Versionen")
        st.caption("Akzeptiert PDF und Excel – jede Kombination moeglich.")

        c_v1, c_v2 = st.columns(2)
        f_old = c_v1.file_uploader(
            "Altes Sortiment (PDF oder Excel)",
            type=["pdf", "xlsx"],
            key="t2_old",
        )
        f_new = c_v2.file_uploader(
            "Neues Sortiment (PDF oder Excel)",
            type=["pdf", "xlsx"],
            key="t2_new",
        )

        if f_old and f_new:
            with st.spinner("Dateien werden eingelesen ..."):
                old_res, old_anom, old_err = quick_parse_file(f_old)
                new_res, new_anom, new_err = quick_parse_file(f_new)

            if old_err:
                st.error("Fehler beim Laden von '" + f_old.name + "': " + old_err)
            if new_err:
                st.error("Fehler beim Laden von '" + f_new.name + "': " + new_err)

            # Anomalie-Warnungen (nicht blockierend)
            if old_anom:
                st.warning(
                    str(old_anom) + " Auffaelligkeit(en) in '" + f_old.name +
                    "'. Fuer eine ausfuehrliche Pruefung bitte in Tab 1 laden."
                )
            if new_anom:
                st.warning(
                    str(new_anom) + " Auffaelligkeit(en) in '" + f_new.name +
                    "'. Fuer eine ausfuehrliche Pruefung bitte in Tab 1 laden."
                )

            if old_res and new_res:
                show_diff_ui(old_res, new_res)
